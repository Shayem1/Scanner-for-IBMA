from tkinter import *
import tkinter as tk
import customtkinter
import time
import sys
import datetime
import xlsxwriter
import os
from openpyxl import Workbook
from openpyxl import load_workbook
# importing packages


# function to save data
def save_data():
    save_data_manual()
    # Set up the next save to occur after 5 minutes
    GUI.after(15000, save_data)

def save_data_manual():
    global workbook, worksheet, wb, sheet, current_date, day_of_week
    current_date = datetime.datetime.now()
    day_of_week = current_date.strftime('%A')
    name = str(day_of_week)+" "+str(day)+"-"+str(month)+"-"+str(T[0])+".xlsx"

    if skip == True:
        try:
            workbook.close()
            if os.path.exists(name):
                os.remove(name)
                workbook = xlsxwriter.Workbook(name)
                worksheet = workbook.add_worksheet()
                worksheet.write(0, 1, "Id")
                worksheet.set_column("A:G", 20)
                workbook.close()
                row = 2
                wb = load_workbook(name, data_only=True)
                sheet = wb.active

                for i in button_list:
                    id = i.cget("text")
                    sheet.cell(row,1).value = str(row)
                    sheet.cell(row,2).value = str(id)
                    row += 1

                wb.template = False
                wb.save(name)
        except:
            print("stawp touching my files :(")
    
# function to allow overwriting of data
def overwrite():
    global skip, workbook, worksheet, row, sheet, name
    # allows it to skip warning
    skip = True

    #removes the warning and sets buttons to active again
    warning_frame.destroy()
    input_id.configure(state=NORMAL)
    save_button.configure(state=NORMAL)
    del_red.configure(state=NORMAL)

    os.remove(name)

    workbook = xlsxwriter.Workbook(name)

    worksheet = workbook.add_worksheet()

    worksheet.write(0, 1, "Id")
    worksheet.set_column("A:G", 20)
    workbook.close()

    row = 2
    wb = load_workbook(name, data_only=True)
    sheet = wb.active

    #removes data


# load data
def merge():
    global skip, workbook, worksheet, row, sheet, name, index, IDlist, StudID, button_list, id,red_list

    warning_frame.destroy()
    input_id.configure(state=NORMAL)
    save_button.configure(state=NORMAL)
    del_red.configure(state=NORMAL)
    skip = True


    wb = load_workbook(name, data_only=True)
    sheet = wb.active

    
    column_data = []
    data_started = False  # Flag to check when valid data starts

    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        col1_value, col2_value = row  # Extract values from column A and B

        if col1_value is not None:
            data_started = True  # Data has started
        elif data_started:  
            break  # Stop when an empty cell appears after valid data
        
        if data_started and col2_value is not None:  # Only append if col2_value is not None
            column_data.append(col2_value)

        index = index + len(column_data)

    os.remove(name)

    workbook = xlsxwriter.Workbook(name)

    worksheet = workbook.add_worksheet()

    worksheet.write(0, 1, "Id")
    worksheet.set_column("A:G", 20)
    workbook.close()

    row = 2
    wb = load_workbook(name, data_only=True)
    sheet = wb.active

    for id in column_data:
        # student ID is sent to be validated
        validation = data_validation(id)

        #adds ID's
        color = validation[1]
        Tcolor = validation[2]
        StudID = customtkinter.CTkButton(master=IDframe, text=id, font=("arial",35), fg_color=color, text_color=Tcolor, corner_radius=10)
        StudID.pack(pady=10)
        
        #makes the scroll wheel to the bottom/new entries
        GUI.update_idletasks()
        canvas.yview_moveto(1)

        # stores ID's in a list
        IDlist.append(id)
        button_list.append(StudID)
        StudID.configure(command=lambda button = StudID: selection(button))

        sheet.cell(index,1).value = str(index)
        sheet.cell(index,2).value = str(id)
        wb.save(name)
        index = index +1

        if validation[0] == False:
            red_list.append(StudID)
    save_data_manual()

# function to remove invalid ID's
def remove_red():
    global red_list, button_list
    # Iterate over a copy of red_list to avoid modifying it while iterating
    for i in red_list:
        i.destroy()
        if i in button_list:
            button_list.remove(i)
        red_list.remove(i)
    if len(red_list) > 0:
        remove_red()
    selected_student.configure(text="Selected Student:\nNo one selected")
    del_selected_ID.configure(state=DISABLED)


# function to remove certain ID
def selected_id(id):
    global button_list, red_list
    id.destroy()
    if id in button_list:
        button_list.remove(id)
    if id in red_list:
        red_list.remove(id)
    del_selected_ID.configure(state=DISABLED)
    selected_student.configure(text="Selected Student:\nNo one selected")


# function to select ID
def selection(id):
    selected_student.configure(text="Selected Student:\n"+str(id.cget("text")))
    del_selected_ID.configure(state=NORMAL)
    del_selected_ID.configure(command=lambda: selected_id(id))
    GUI.update_idletasks()

# scanner setup
def data_validation(string):
    try:
        if len(string) == 8 and int(string[0:4]) > 2000 and int(string[0:4]) < 2100 and string.isnumeric() == True:
            list1 = [True,"#13005e", "White"]
            
        elif len(string) == 9 and str(string[0:3])=="IBM" and string[3:9].isnumeric() == True:
            list1 = [True, "#0f39d4", "Black"]

        else:
            list1 = [False, "red", "Black"]
    except:
        list1 = [False, "red", "Black"]

    return list1

# scanner main entry
def ID_entered(event):
    global IDlist, StudID, button_list, id, index, wb, sheet, red_list
    # student ID is sent to be validated
    id = input_id.get()
    validation = data_validation(id)
    # resets the entry field
    input_id.delete(0, END)

    #adds ID's
    color = validation[1]
    Tcolor = validation[2]
    StudID = customtkinter.CTkButton(master=IDframe, text=id, font=("arial",35), fg_color=color, text_color=Tcolor, corner_radius=10)
    StudID.pack(pady=10)
    
    #makes the scroll wheel to the bottom/new entries
    GUI.update_idletasks()
    canvas.yview_moveto(1)

    # stores ID's in a list
    IDlist.append(id)
    button_list.append(StudID)
    StudID.configure(command=lambda button = StudID: selection(button))
    
    

    sheet.cell(index,1).value = str(index)
    sheet.cell(index,2).value = str(id)
    wb.save(name)
    index = index +1

    if validation[0] == False:
        red_list.append(StudID)

    time_update_manual()

def time_update_manual():
    global current_date, day_of_week, T, day, month
    current_date = datetime.datetime.now()
    day_of_week = current_date.strftime('%A')
    T = time.localtime()
    minute = T[4]
    day = T[2]
    month = T[1]
    if len(str(minute)) == 1:
        minute = "0"+str(minute)
    if len(str(day)) == 1:
        day = "0"+str(day)
    if len(str(month)) == 1:
        month = "0"+str(month)
    today = "Today is:\n\n"+str(day_of_week)+" "+str(day)+"-"+str(month)+"-"+str(T[0])+"  "+str(T[3])+":"+str(minute)
    time_text.configure(text=today)

def time_update():
    global current_date, day_of_week, T, day, month
    current_date = datetime.datetime.now()
    day_of_week = current_date.strftime('%A')
    T = time.localtime()
    minute = T[4]
    day = T[2]
    month = T[1]
    if len(str(minute)) == 1:
        minute = "0"+str(minute)
    if len(str(day)) == 1:
        day = "0"+str(day)
    if len(str(month)) == 1:
        month = "0"+str(month)
    today = "Today is:\n\n"+str(day_of_week)+" "+str(day)+"-"+str(month)+"-"+str(T[0])+"  "+str(T[3])+":"+str(minute)
    time_text.configure(text=today)
    GUI.after(10000, time_update)
    

# initialising the GUI
GUI = customtkinter.CTk()
GUI.geometry("1200x800")
GUI.maxsize(width=1200, height=800)
GUI.minsize(width=1200, height=800)
GUI.title("Student ID Scanner")
GUI.grid_rowconfigure(1, weight=1, uniform="equal")
GUI.grid_columnconfigure(2, weight=1, uniform="equal")
customtkinter.set_appearance_mode("dark")

global skip
IDlist = [] #temp variable
button_list = []
red_list = []
index = 2
skip = True

# initialising student ID output frame
IDframe = customtkinter.CTkScrollableFrame(master= GUI, width= 300, corner_radius=30)
IDframe.grid(padx = 50, row=1, column = 0,  sticky="nsew", pady=20)
canvas = IDframe._parent_canvas

option_frame = customtkinter.CTkFrame(master=GUI, width=670, height= 600, corner_radius=20)
option_frame.grid(padx = 25, row=1, column = 1,  sticky="nsew", pady=20)

# time setup
current_date = datetime.datetime.now()
day_of_week = current_date.strftime('%A')
T = time.localtime()
minute = T[4]
day = T[2]
month = T[1]
if len(str(minute)) == 1:
    minute = "0"+str(minute)
if len(str(day)) == 1:
        day = "0"+str(day)
if len(str(month)) == 1:
    month = "0"+str(month)
time_text = customtkinter.CTkLabel(master= option_frame, font=("arial",40))
time_text.place(relx = 0.5, rely = 0.05, anchor = N)

#save data setup
name = str(day_of_week)+" "+str(day)+"-"+str(month)+"-"+str(T[0])+".xlsx"

# text setup
studentid_text = customtkinter.CTkLabel(master= GUI, text="   Student ID   ", font=("arial",45), fg_color="#41229c", corner_radius=10)
studentid_text.grid(pady=15, row=0, column = 0)

OPTIONS_text = customtkinter.CTkLabel(master= GUI, text="   OPTIONS   ", font=("arial",45), fg_color="#41229c", corner_radius=10)
OPTIONS_text.grid(pady=15, row=0, column = 1, columnspan = 4)

manual_text = customtkinter.CTkLabel(master= option_frame, text="Manual Entry", font=("arial",40), fg_color="#41229c", corner_radius=10)
manual_text.place(relx = 0.05, rely =0.35)

del_red = customtkinter.CTkButton(master= option_frame, text="Delete All Red", font=("arial",40), fg_color="#f50a0e", corner_radius=15, height=60)
del_red.configure( command= lambda: remove_red())
del_red.place(relx = 0.5, rely =0.85)

del_selected_ID = customtkinter.CTkButton(master= option_frame, text="Delete \n selected ID", state=DISABLED,font=("arial",40), fg_color="#f50a0e", corner_radius=10)
del_selected_ID.place(relx = 0.07, rely =0.55)

selected_student = customtkinter.CTkLabel(master=option_frame, text="Selected Student:\nNo one selected", font=("arial",35))
selected_student.place(relx = 0.52, rely = 0.56)

save_button = customtkinter.CTkButton(master= option_frame, text="SAVE", state=NORMAL,font=("arial",40), fg_color="#30cf0c", corner_radius=10)
save_button.place(relx = 0.13, rely =0.86)
save_button.configure(command = lambda: save_data_manual())

#set up entry box
input_id = customtkinter.CTkEntry(master=option_frame, placeholder_text = "Write ID here", font=("arial",35), width=300, height= 45)
input_id.place(relx = 0.5, rely = 0.35)

# pressing enter will accept the entry
input_id.bind("<Return>", ID_entered)

#binds the curser to the entry field
input_id.focus_force()

#prevents overwriting data after a crash
if os.path.exists(name):
    global choice
    skip = False
    # disabled buttons
    input_id.configure(state=DISABLED)
    save_button.configure(state=DISABLED)
    del_red.configure(state=DISABLED)

    choice = tk.StringVar()

    
    warning_frame = customtkinter.CTkFrame(master = GUI,width = 600, height = 250, corner_radius = 20, fg_color="grey92", bg_color="grey18")
    warning_frame.place(relx = 0.5, rely = 0.5, anchor = CENTER)

    #information
    warning_label = customtkinter.CTkLabel(master=warning_frame, text="WARNING - DATA FOUND", bg_color="grey92", text_color="red", font=("arial",40))
    warning_label.place(relx = 0.5, rely = 0.2, anchor= CENTER)

    merge_info = customtkinter.CTkLabel(master=warning_frame, text="Merge: Adds new data onto old file", bg_color="grey92", text_color="Black", font=("arial",20))
    merge_info.place(relx = 0.5, rely = 0.42, anchor= CENTER)

    overwrite_info = customtkinter.CTkLabel(master=warning_frame, text="Overwrite: Removes old data and creates a new file", bg_color="grey92", text_color="Black", font=("arial",20))
    overwrite_info.place(relx = 0.5, rely = 0.57, anchor= CENTER)
    

    #options for data
    merge_button = customtkinter.CTkButton(master=warning_frame, text="Merge", corner_radius=15, font=("arial",25),command= lambda: merge())
    merge_button.place(relx = 0.3, rely = 0.8, anchor = CENTER)

    overwrite_button = customtkinter.CTkButton(master=warning_frame, text="Overwrite", corner_radius=15, font=("arial",25), command= lambda: overwrite())
    overwrite_button.place(relx = 0.7, rely = 0.8, anchor = CENTER)

if skip == True:
    workbook = xlsxwriter.Workbook(name)

    worksheet = workbook.add_worksheet()

    worksheet.write(0, 1, "Id")
    worksheet.set_column("A:G", 20)
    workbook.close()

    row = 2
    wb = load_workbook(name, data_only=True)
    sheet = wb.active

time_update()
save_data()


GUI.mainloop()